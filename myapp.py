import pandas as pd
import numpy as np
import streamlit as st
import io
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font, PatternFill
import sys
from openpyxl.utils import get_column_letter
st.title("Roster Comparison Report")
uploaded_file = st.file_uploader("Upload Excel File", type=["xlsx"])
if uploaded_file is not None:
        
    df = pd.read_excel(uploaded_file, engine='openpyxl',sheet_name='Sheet1')
    #st.title("Analysis report of AI 16-30JUNE_CC 1")

    #st.write("Data Preview:", df.head())

    total_crew = (df['Staff ID'].unique())
    total_assignments = df['Staff ID'].count()
    req_rows = 200000
    req_cols = 35
    new_df = pd.DataFrame(index=range(req_rows), columns=range(req_cols))

    new_df.iloc[0, 0] = "Total Crew"
    new_df.iloc[0, 1] = len(total_crew) 
    new_df.iloc[1, 0] = "Total Assignments"
    new_df.iloc[1, 1] = total_assignments

    st.write("Total Crew:", len(total_crew))
    st.write("Total Assignments:", total_assignments)
    st.write("Data Preview:", df.head())

    # for activity type change table
    start_row = 3
    start_col = 0
    activity_types = df['Activity Type Change'].unique()
    for i, activity in enumerate(activity_types): 
        new_df.iloc[start_row + i, start_col] = activity
        new_df.iloc[start_row + i, start_col + 1] = df[df['Activity Type Change'] == activity].shape[0]
        new_df.iloc[start_row + i, start_col + 2] = str(((df[df['Activity Type Change'] == activity].shape[0]/total_assignments) * 100).round(2)) + "%"

    changed = df[(df['Activity Type Change']=='Changes Made') & (df["Activity Type_published"] != df["Activity Type_ingested"])]

    pivot = pd.pivot_table(
        changed,
        index="Activity Type_published",
        columns="Activity Type_ingested",
        values="Staff ID",
        aggfunc="count",
        fill_value="",
        margins=True,
        margins_name="TOTAL"
    )
    new_df.loc[11,0] = "CHANGES MADE"
    pivot_start_row = 12
    pivot_start_col = 0

    pivot_df = pivot.reset_index()

    # Add column headers for the pivot table
    for j, col_name in enumerate(pivot_df.columns):
        new_df.iloc[pivot_start_row, pivot_start_col + j] = col_name

    # Write pivot data starting from the next row
    for i, row in pivot_df.iterrows():
        for j, value in enumerate(row):
            new_df.iloc[pivot_start_row + 1 + i, pivot_start_col + j] = value

    deleted = df[df['Activity Type Change'] == 'Deleted']
    deleted_counts = deleted['Activity Type_ingested'].value_counts().reset_index()
    deleted_counts.columns = ['ACTIVITY CODE', 'DELETED COUNT']

    new_activity = df[df['Activity Type Change'] == 'New']
    new_counts = new_activity['Activity Type_published'].value_counts().reset_index()
    new_counts.columns = ['ACTIVITY CODE', 'NEW COUNT']

    merged = pd.merge(deleted_counts, new_counts, on='ACTIVITY CODE', how='outer').fillna(0)

    merged['DELETED COUNT'] = merged['DELETED COUNT'].astype(int)
    merged['NEW COUNT'] = merged['NEW COUNT'].astype(int)
    merged['DIFFERENCE'] = merged['NEW COUNT'] - merged['DELETED COUNT']

    start_row = pivot_start_row + len(pivot_df) + 4

    new_df.iloc[start_row - 1, 0] = "DELETED ACTIVITY CODE"
    new_df.iloc[start_row - 1, 1] = "COUNT"
    new_df.iloc[start_row - 1, 2] = "NEW ACTIVITY CODE"
    new_df.iloc[start_row - 1, 3] = "COUNT"
    new_df.iloc[start_row - 1, 4] = "DIFFERENCE"

    for i, row in merged.iterrows():
        new_df.iloc[start_row + i, 0] = row['ACTIVITY CODE']
        new_df.iloc[start_row + i, 1] = row['DELETED COUNT']
        new_df.iloc[start_row + i, 2] = row['ACTIVITY CODE']
        new_df.iloc[start_row + i, 3] = row['NEW COUNT']
        new_df.iloc[start_row + i, 4] = row['DIFFERENCE']

    # --- FLY Removals summary and crew breakdown ---

    # Get FLY removals per staff
    deleted_fly = df[(df['Activity Type Change'] == 'Deleted') & (df['Activity Type_ingested'] == 'FLY')]
    fly_removals = deleted_fly['Staff ID'].value_counts().reset_index()
    fly_removals.columns = ['Staff ID', 'Number of FLY Removals']

    # Categorize
    def fly_category(n):
        if n >= 10:
            return ">=10"
        elif n >= 5:
            return ">=5, <=9"
        elif n >= 1:
            return ">=1, <=4"
        else:
            return ""

    fly_removals['Category'] = fly_removals['Number of FLY Removals'].apply(fly_category)

    # Summary table
    summary = [
        [">=10", fly_removals[fly_removals['Category'] == ">=10"]['Number of FLY Removals'].sum(), fly_removals[fly_removals['Category'] == ">=10"].shape[0]],
        [">=5, <=9", fly_removals[fly_removals['Category'] == ">=5, <=9"]['Number of FLY Removals'].sum(), fly_removals[fly_removals['Category'] == ">=5, <=9"].shape[0]],
        [">=1, <=4", fly_removals[fly_removals['Category'] == ">=1, <=4"]['Number of FLY Removals'].sum(), fly_removals[fly_removals['Category'] == ">=1, <=4"].shape[0]],
        ["Grand Total", fly_removals['Number of FLY Removals'].sum(), fly_removals.shape[0]]
    ]

    # Find where to start the FLY table (right of merged, e.g. col 7)
    fly_start_row = start_row - 1
    fly_start_col = 7

    # Write FLY summary headers
    new_df.iloc[fly_start_row, fly_start_col] = "Categories"
    new_df.iloc[fly_start_row, fly_start_col + 1] = "Number of FLY Removals"
    new_df.iloc[fly_start_row, fly_start_col + 2] = "Crew Count"

    # Write FLY summary data
    for i, row in enumerate(summary):
        for j, value in enumerate(row):
            new_df.iloc[fly_start_row + 1 + i, fly_start_col + j] = value

    # Write detailed breakdown headers
    detail_start_row = fly_start_row + len(summary) + 2
    new_df.iloc[detail_start_row, fly_start_col] = "Staff ID"
    new_df.iloc[detail_start_row, fly_start_col + 1] = "Number of FLY Removals"
    new_df.iloc[detail_start_row, fly_start_col + 2] = "Category"

    fly_removals_gt10 = fly_removals[fly_removals['Category'] == ">=10"]
    for i, row in fly_removals_gt10.iterrows():
        new_df.iloc[detail_start_row + 1 + i, fly_start_col] = row['Staff ID']
        new_df.iloc[detail_start_row + 1 + i, fly_start_col + 1] = row['Number of FLY Removals']
        new_df.iloc[detail_start_row + 1 + i, fly_start_col + 2] = row['Category']

    added_fly = df[(df['Activity Type Change'] == 'New') & (df['Activity Type_published'] == 'FLY')]
    fly_added = added_fly['Staff ID'].value_counts().reset_index()
    fly_added.columns = ['Staff ID', 'Number of FLY Additions']

    fly_added['Category'] = fly_added['Number of FLY Additions'].apply(fly_category)
    # Summary table for added FLY
    added_summary = [
        [">=10", fly_added[fly_added['Category'] == ">=10"]['Number of FLY Additions'].sum(), fly_added[fly_added['Category'] == ">=10"].shape[0]],
        [">=5, <=9", fly_added[fly_added['Category'] == ">=5, <=9"]['Number of FLY Additions'].sum(), fly_added[fly_added['Category'] == ">=5, <=9"].shape[0]],
        [">=1, <=4", fly_added[fly_added['Category'] == ">=1, <=4"]['Number of FLY Additions'].sum(), fly_added[fly_added['Category'] == ">=1, <=4"].shape[0]],
        ["Grand Total", fly_added['Number of FLY Additions'].sum(), fly_added.shape[0]]
    ]
    fly_start_row_added = detail_start_row + len(fly_removals_gt10) + 3
    fly_start_col_added = fly_start_col 
    # Write added FLY summary headers
    new_df.iloc[fly_start_row_added, fly_start_col_added] = "Categories"
    new_df.iloc[fly_start_row_added, fly_start_col_added + 1] = "Number of FLY Additions"
    new_df.iloc[fly_start_row_added, fly_start_col_added + 2] = "Crew Count"

    # Write added FLY summary data
    for i, row in enumerate(added_summary):
        for j, value in enumerate(row):
            new_df.iloc[fly_start_row_added + 1 + i, fly_start_col_added + j] = value
            
    # Write detailed breakdown headers for added FLY
    detail_start_row_added = fly_start_row_added + len(added_summary) + 2
    new_df.iloc[detail_start_row_added, fly_start_col_added] = "Staff ID"
    new_df.iloc[detail_start_row_added, fly_start_col_added + 1] = "Number of FLY Additions"
    new_df.iloc[detail_start_row_added, fly_start_col_added + 2] = "Category"

    fly_added_gt10 = fly_added[fly_added['Category'] == ">=10"]
    for i, row in fly_added_gt10.iterrows():
        new_df.iloc[detail_start_row_added + 1 + i, fly_start_col_added] = row['Staff ID']
        new_df.iloc[detail_start_row_added + 1 + i, fly_start_col_added + 1] = row['Number of FLY Additions']
        new_df.iloc[detail_start_row_added + 1 + i, fly_start_col_added + 2] = row['Category']



    # Save your DataFrame to Excel first
    output = io.BytesIO()
    new_df.to_excel(output, index=False, header=False, engine='openpyxl')
    output.seek(0)

    # Load workbook and select the first worksheet
    wb = load_workbook(output)
    ws = wb.active

    # Define border style
    thin = Side(border_style="thin", color="000000")
    thick = Side(border_style="thick", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)  # <-- Add this line

    # Apply border to all non-empty cells
    for row in ws.iter_rows():
        for cell in row:
            if cell.value not in (None, ""):
                cell.border = border

    cell = ws["A1:A2"]
    for row in ws["A1:A2"]:
        for c in row:
            c.font = Font(bold=True)
            c.fill = PatternFill(start_color="D9CAE3", end_color="D9CAE3", fill_type="solid")  # Light purple

    cell = ws.cell(row=12, column=1)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="FFDAB9", end_color="FFDAB9", fill_type="solid")

    pivot_height = len(pivot_df) + 1  # +1 for header
    pivot_width = len(pivot_df.columns)

    for i in range(pivot_height):
        for j in range(pivot_width):
            cell = ws.cell(row=pivot_start_row + i, column=pivot_start_col + j + 1)
            # Set thin borders by default
            left   = thin
            right  = thin
            top    = thin
            bottom = thin
            # Set thick border on the outside
            if i == 0:
                top = thick
            if i == pivot_height - 1:
                bottom = thick
            if j == 0:
                left = thick
            if j == pivot_width - 1:
                right = thick
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)

    ouput2 = io.BytesIO()
    wb.save(ouput2)
    ouput2.seek(0)
    st.download_button(
        label="Download Analysis Report",
        data=ouput2,
        file_name="Roster Comparision report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


