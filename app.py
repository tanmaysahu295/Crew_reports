import pandas as pd
import numpy as np
import streamlit as st
import openpyxl
import io
from openpyxl.styles import PatternFill

st.title("Crew Non-Availability Report")
uploaded_file = st.file_uploader("Upload Excel File", type=["xlsx"])

if uploaded_file is not None:
    df = pd.read_excel(uploaded_file, engine='openpyxl')
    st.write("Data Preview:", df.head())
    filtered_df = df[df['Mapped Code'].isnull()]
    not_filtered_df = df[df['Mapped Code'].notnull()]

    filtered_df.loc[filtered_df["ACTIVITY TYPE "].str.contains("CONVERSION",na=False), "Mapped Code"] = "CONVERSION"
    filtered_df.loc[filtered_df["ACTIVITY TYPE "] == "MISC_MT", "Mapped Code"] = "MISSCARRAIGE"
    filtered_df.loc[filtered_df["ACTIVITY TYPE "] == "MISSED BA", "Mapped Code"] = "BA +ve"
    filtered_df.loc[filtered_df["ACTIVITY TYPE "].str.contains('CRM|DRILL|REF|IND_TRG|DGR|AGTR|POLAR|PACIFICBFG|CCQ|UPRT_GRD'), "Mapped Code"] = "GROUND TRG"
    filtered_df.loc[filtered_df["ACTIVITY TYPE "].str.contains("ML",na= False), "Mapped Code"] = "ML/PREGNANCY"
    filtered_df.loc[filtered_df["ACTIVITY TYPE "].str.contains("SBY|HLBY",na= False), "Mapped Code"] = "STANDBY"
    filtered_df.loc[filtered_df["ACTIVITY TYPE "] == "NAT", "Mapped Code"] = "NAT"
    filtered_df.loc[filtered_df["ACTIVITY TYPE "] == "OJI", "Mapped Code"] = "OJI"
    filtered_df.loc[filtered_df["ACTIVITY TYPE "].str.contains("TMU",na= False), "Mapped Code"] = "TMU"
    filtered_df.loc[filtered_df["ACTIVITY TYPE "].str.contains("PMU",na= False), "Mapped Code"] = "PMU"
    filtered_df.loc[filtered_df["ACTIVITY TYPE "].str.contains("PL",na= False), "Mapped Code"] = "PL"
    filtered_df.loc[filtered_df["ACTIVITY TYPE "].str.contains("CL",na= False), "Mapped Code"] = "CL"
    filtered_df.loc[filtered_df["ACTIVITY TYPE "].str.contains("SL",na= False), "Mapped Code"] = "SL"
    filtered_df.loc[filtered_df["ACTIVITY TYPE "] == "Blank", "Mapped Code"] = "REST"
    filtered_df.loc[filtered_df["ACTIVITY TYPE "] == "RLL", "Mapped Code"] = "RELOCATION"
    filtered_df.loc[filtered_df["ACTIVITY TYPE "] == "EWLB", "Mapped Code"] = "LEAVE BLOCK"
    filtered_df.loc[filtered_df["ACTIVITY TYPE "] == "FAT", "Mapped Code"] = "REST"
    filtered_df.loc[filtered_df["ACTIVITY TYPE "] == "NME", "Mapped Code"] = "MEDICAL NOT DONE FROM CREW'S END"
    filtered_df.loc[filtered_df["ACTIVITY TYPE "].str.contains("OFF",na= False), "Mapped Code"] = "REST"
    filtered_df.loc[filtered_df["ACTIVITY TYPE "].str.contains("SIM",na= False), "Mapped Code"] = "SIM TRAING"

    st.write("Filtered Data Preview:", filtered_df.head())
    st.write("Not Filtered Data Preview:", not_filtered_df.head())

    combined_df = pd.concat([filtered_df, not_filtered_df], ignore_index=True)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        combined_df.to_excel(writer, index=False, sheet_name='Sheet1')
        worksheet = writer.sheets['Sheet1']

        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = length + 2

        
        header_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid") 
        for cell in next(worksheet.iter_rows(min_row=1, max_row=1)):
            cell.fill = header_fill

    output.seek(0)
    st.download_button(
        label="Download Filtered Data",
        data=output,
        file_name="filtered_data.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )